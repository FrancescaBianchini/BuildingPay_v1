# -*- coding: utf-8 -*-
import uuid
import logging
from datetime import date, timedelta
from odoo import api, fields, models, _
from odoo.exceptions import ValidationError, UserError

_logger = logging.getLogger(__name__)


class ResPartner(models.Model):
    """
    Estensione del modello res.partner per BuildingPay.
    Aggiunge:
    - Tipo contatto 'Condominio'
    - Flag Amministratore
    - Gestione contratto e flag accordo
    - Referrer
    - Listino condominio
    - Campi specifici: fiscalcode, pec_mail, codice_destinatario, IBAN
    """
    _inherit = 'res.partner'

    # -------------------------------------------------------
    # Estensione tipo contatto con 'condominio'
    # -------------------------------------------------------
    type = fields.Selection(
        selection_add=[('condominio', 'Condominio')],
        ondelete={'condominio': 'set default'},
    )

    # -------------------------------------------------------
    # Flag Amministratore (gestito dal gruppo BuildingPay Manager)
    # -------------------------------------------------------
    is_amministratore = fields.Boolean(
        string='Amministratore',
        default=False,
        tracking=True,
        help='Indica se il contatto è un amministratore di condomini.',
    )
    referrer_code = fields.Char(
        string='Codice Referrer',
        readonly=True,
        copy=False,
        index=True,
        help='Codice univoco per invitare nuovi amministratori tramite link referral.',
    )
    referrer_id = fields.Many2one(
        comodel_name='res.partner',
        string='Referrer',
        ondelete='set null',
        tracking=True,
        help='Contatto che ha invitato questo amministratore.',
    )
    referred_ids = fields.One2many(
        comodel_name='res.partner',
        inverse_name='referrer_id',
        string='Amministratori invitati',
    )

    # -------------------------------------------------------
    # Listino condominio associato (solo per amministratori)
    # -------------------------------------------------------
    listino_condominio_id = fields.Many2one(
        comodel_name='product.pricelist',
        string='Listino Condominio',
        domain=[('is_listino_condominio', '=', True)],
        help='Listino prezzi condominio associato a questo amministratore.',
    )

    # -------------------------------------------------------
    # Contratto / Accordo Condomini Aggregati
    # -------------------------------------------------------
    accordo_condomini_aggregati = fields.Boolean(
        string='Accordo condomini aggregati',
        default=False,
        tracking=True,
        help='Attivato quando l\'amministratore carica l\'Accordo Condomini Aggregati firmato.',
    )
    contratto_file = fields.Binary(
        string='File Accordo Condomini Aggregati',
        attachment=True,
    )
    contratto_filename = fields.Char(
        string='Nome file contratto',
    )
    contratto_upload_date = fields.Datetime(
        string='Data caricamento contratto',
        readonly=True,
    )

    # -------------------------------------------------------
    # Data archiviazione (per indirizzi di tipo condominio)
    # -------------------------------------------------------
    data_archiviazione = fields.Date(
        string='Data archiviazione',
        readonly=True,
        help='Data in cui l\'indirizzo di tipo condominio è stato archiviato.',
    )

    # -------------------------------------------------------
    # Campi italiani specifici (compatibili con l10n_it_edi)
    # Definiti solo se non già presenti tramite l10n_it
    # -------------------------------------------------------
    # Nota: fiscalcode, pec_mail, codice_destinatario potrebbero
    # essere già definiti da l10n_it_edi. In quel caso usare
    # 'related' o rinominare nel caso si trovino con nomi diversi.
    # Qui vengono aggiunti come campi custom se non già esistenti.

    # fiscalcode: campo codice fiscale (l10n_it_edi lo definisce come
    # 'l10n_it_codice_fiscale'; se si usa l10n_it_edi commentare qui
    # e usare il campo nativo)
    fiscalcode = fields.Char(
        string='Codice Fiscale',
        size=16,
        index=True,
    )
    # pec_mail: email PEC (l10n_it_edi usa 'l10n_it_pec_email')
    pec_mail = fields.Char(
        string='Email PEC',
        help='Indirizzo email PEC del contatto.',
    )
    # codice_destinatario: codice SDI per fatturazione elettronica
    # (l10n_it_edi usa 'l10n_it_pa_index')
    codice_destinatario = fields.Char(
        string='Codice Destinatario SDI',
        size=7,
        help='Codice destinatario (SDI) per la fatturazione elettronica.',
    )
    electronic_invoice_subjected = fields.Boolean(
        string='Soggetto a fatturazione elettronica',
        default=False,
    )
    electronic_invoice_obliged_subject = fields.Boolean(
        string='Obbligo fatturazione elettronica',
        default=False,
    )

    # -------------------------------------------------------
    # Condomini figli (indirizzi di tipo condominio)
    # -------------------------------------------------------
    condominio_ids = fields.One2many(
        comodel_name='res.partner',
        inverse_name='parent_id',
        string='Condomini',
        domain=[('type', '=', 'condominio'), ('active', '=', True)],
    )
    condominio_count = fields.Integer(
        string='Numero Condomini',
        compute='_compute_condominio_count',
    )

    # -------------------------------------------------------
    # Metodi compute
    # -------------------------------------------------------
    @api.depends('child_ids', 'child_ids.type', 'child_ids.active')
    def _compute_condominio_count(self):
        for partner in self:
            partner.condominio_count = self.env['res.partner'].search_count([
                ('parent_id', '=', partner.id),
                ('type', '=', 'condominio'),
                ('active', '=', True),
            ])

    # -------------------------------------------------------
    # Onchange e constraints
    # -------------------------------------------------------
    @api.model_create_multi
    def create(self, vals_list):
        partners = super().create(vals_list)
        for partner in partners:
            if partner.is_amministratore and not partner.referrer_code:
                partner.referrer_code = self._generate_referrer_code()
        return partners

    def write(self, vals):
        result = super().write(vals)
        for partner in self:
            if partner.is_amministratore and not partner.referrer_code:
                partner.referrer_code = self._generate_referrer_code()
        return result

    @api.model
    def _generate_referrer_code(self):
        """Genera un codice referrer univoco."""
        return uuid.uuid4().hex[:12].upper()

    # -------------------------------------------------------
    # Upload contratto
    # -------------------------------------------------------
    def action_upload_contratto(self, file_data, filename):
        """
        Chiamato dal portale quando l'utente carica l'accordo firmato.
        Attiva il flag accordo_condomini_aggregati e crea eventuale attività.
        """
        self.ensure_one()
        self.write({
            'contratto_file': file_data,
            'contratto_filename': filename,
            'accordo_condomini_aggregati': True,
            'contratto_upload_date': fields.Datetime.now(),
        })
        # Crea attività automatica se configurato
        self._create_contratto_activity()

    def _create_contratto_activity(self):
        """Crea attività automatica se la configurazione lo richiede."""
        self.ensure_one()
        config = self.env['buildingpay.config'].get_config_for_website()
        if not config or not config.create_activity_on_contract:
            return
        if not config.activity_responsible_id:
            return

        deadline = date.today() + timedelta(days=config.activity_days or 5)
        activity_type = self.env.ref('mail.mail_activity_data_todo', raise_if_not_found=False)

        self.activity_schedule(
            activity_type_id=activity_type.id if activity_type else False,
            summary=_('Controllare il contratto Accordo Condomini Aggregati '
                       'caricato dall\'amministratore'),
            date_deadline=deadline,
            user_id=config.activity_responsible_id.id,
        )

    # -------------------------------------------------------
    # Archiviazione indirizzo condominio
    # -------------------------------------------------------
    def action_archive_condominio(self):
        """
        Archivia un indirizzo di tipo condominio.
        Salva la data di archiviazione e genera Excel notifica.
        """
        self.ensure_one()
        if self.type != 'condominio':
            raise UserError(_('Solo gli indirizzi di tipo "Condominio" possono essere archiviati.'))

        self.write({
            'data_archiviazione': fields.Date.today(),
            'active': False,
        })
        # Genera e invia Excel notifica condominio dismesso
        self._send_condominio_dismesso_email()

    def _send_condominio_dismesso_email(self):
        """Genera il file Excel del condominio dismesso e lo invia per email."""
        self.ensure_one()
        config = self.env['buildingpay.config'].get_config_for_website()
        if not config or not config.condomini_dismessi_email:
            return

        try:
            import openpyxl
            from io import BytesIO
            import base64

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Condominio Dismesso'

            # Intestazioni
            headers = [
                'ID Esterno Amministratore',
                'Nome e Cognome Amministratore',
                'ID Esterno Condominio',
                'Nome Condominio',
                'Indirizzo Completo',
                'Dismesso',
            ]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Dati
            admin = self.parent_id
            admin_ext_id = self.env['ir.model.data'].search([
                ('model', '=', 'res.partner'),
                ('res_id', '=', admin.id if admin else 0),
            ], limit=1)
            cond_ext_id = self.env['ir.model.data'].search([
                ('model', '=', 'res.partner'),
                ('res_id', '=', self.id),
            ], limit=1)

            address = ' '.join(filter(None, [
                self.street, self.zip, self.city,
                self.state_id.name if self.state_id else '',
                self.country_id.name if self.country_id else '',
            ]))
            row_data = [
                admin_ext_id.complete_name if admin_ext_id else '',
                admin.name if admin else '',
                cond_ext_id.complete_name if cond_ext_id else '',
                self.name or '',
                address,
                True,
            ]
            for col, value in enumerate(row_data, 1):
                ws.cell(row=2, column=col, value=value)

            output = BytesIO()
            wb.save(output)
            excel_data = base64.b64encode(output.getvalue())

            # Invio email
            recipients = [e.strip() for e in config.condomini_dismessi_email.split(',') if e.strip()]
            if recipients:
                mail_values = {
                    'subject': _('Condominio dismesso: %s') % self.name,
                    'body_html': _(
                        '<p>Il condominio <strong>%s</strong> è stato dismesso in data %s.</p>'
                        '<p>In allegato il file Excel con i dettagli.</p>'
                    ) % (self.name, fields.Date.today()),
                    'email_to': ','.join(recipients),
                    'attachment_ids': [(0, 0, {
                        'name': 'condominio_dismesso_%s.xlsx' % self.name,
                        'datas': excel_data,
                        'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    })],
                }
                self.env['mail.mail'].sudo().create(mail_values).send()
        except Exception as e:
            _logger.error('Errore invio email condominio dismesso: %s', e)

    # -------------------------------------------------------
    # Azione pianificata: report giornaliero condomini
    # -------------------------------------------------------
    @api.model
    def action_send_daily_condomini_report(self):
        """
        Azione pianificata: genera e invia il report Excel giornaliero
        di tutti i condomini attivi con i loro amministratori.
        Eseguita ogni giorno alle 23:30.
        """
        try:
            import openpyxl
            from io import BytesIO
            import base64

            # Recupera tutti i condomini attivi
            condominii = self.search([
                ('type', '=', 'condominio'),
                ('active', '=', True),
                ('parent_id', '!=', False),
                ('parent_id.is_amministratore', '=', True),
            ])

            if not condominii:
                _logger.info('BuildingPay: nessun condominio attivo trovato.')
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Condomini Attivi'

            # Intestazioni
            headers = [
                'ID Esterno Amministratore',
                'Nome e Cognome Amministratore',
                'ID Esterno Condominio',
                'Nome Condominio',
                'Indirizzo Completo',
                'IBAN',
                'Email PEC',
                'Codice Destinatario',
                'Codice Fiscale',
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)

            # Dati
            for row_idx, condo in enumerate(condominii, 2):
                admin = condo.parent_id

                # External IDs
                admin_ext_id_rec = self.env['ir.model.data'].search([
                    ('model', '=', 'res.partner'),
                    ('res_id', '=', admin.id),
                ], limit=1)
                condo_ext_id_rec = self.env['ir.model.data'].search([
                    ('model', '=', 'res.partner'),
                    ('res_id', '=', condo.id),
                ], limit=1)

                # IBAN del condominio
                bank = self.env['res.partner.bank'].search([
                    ('partner_id', '=', condo.id),
                ], limit=1)

                address = ' '.join(filter(None, [
                    condo.street,
                    condo.zip,
                    condo.city,
                    condo.state_id.name if condo.state_id else '',
                    condo.country_id.name if condo.country_id else '',
                ]))

                row_data = [
                    admin_ext_id_rec.complete_name if admin_ext_id_rec else '',
                    admin.name or '',
                    condo_ext_id_rec.complete_name if condo_ext_id_rec else '',
                    condo.name or '',
                    address,
                    bank.acc_number if bank else '',
                    condo.pec_mail or '',
                    condo.codice_destinatario or '',
                    condo.fiscalcode or '',
                ]
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col, value=value)

            output = BytesIO()
            wb.save(output)
            excel_data = base64.b64encode(output.getvalue())

            # Recupera destinatari da tutte le configurazioni BuildingPay
            configs = self.env['buildingpay.config'].search([
                ('condomini_attivati_email', '!=', False),
            ])
            all_recipients = set()
            for cfg in configs:
                for email in cfg.condomini_attivati_email.split(','):
                    email = email.strip()
                    if email:
                        all_recipients.add(email)

            if all_recipients:
                today_str = fields.Date.today().strftime('%Y-%m-%d')
                mail_values = {
                    'subject': _('Report Condomini Attivi - %s') % today_str,
                    'body_html': _(
                        '<p>In allegato il report giornaliero dei condomini attivi '
                        'generato in data %s.</p>'
                    ) % today_str,
                    'email_to': ','.join(all_recipients),
                    'attachment_ids': [(0, 0, {
                        'name': 'condomini_attivi_%s.xlsx' % today_str,
                        'datas': excel_data,
                        'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    })],
                }
                self.env['mail.mail'].sudo().create(mail_values).send()
                _logger.info('BuildingPay: report giornaliero inviato a %s', all_recipients)
            else:
                _logger.warning('BuildingPay: nessun destinatario configurato per il report condomini.')

        except Exception as e:
            _logger.error('BuildingPay: errore generazione report giornaliero condomini: %s', e)
